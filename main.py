import csv
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import os
from collections import defaultdict
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo



input_path = os.path.join(os.path.dirname(__file__), 'sales_data.csv')

output_path = os.path.join(os.path.dirname(__file__), 'sales_output_v2.xlsx')

wb = Workbook() #skapar ny Excel fil
ws = wb.active 
ws.title = "Sales" 

ws.append(["Date", "Product", "Region", "Count sold", "Price per item"])

for cell in ws[1]:
    cell.font = Font(bold=True, size=14)

with open(input_path, newline="", encoding="utf-8") as file:
    reader = csv.reader(file)
    next(reader)

    for row in reader:
        ws.append(row)

    ws.insert_cols(6)
    ws.cell(row=1, column=6, value="Total sales").font = Font(bold=True, size=14)


for row in range(2, ws.max_row + 1): 
    count = float(ws.cell(row=row, column=4).value)
    price = float(ws.cell(row=row, column=5).value)

    ws.cell(row=row, column=6).value = count * price


region_ws = wb.create_sheet("Region Sales")


region_ws.append(["Region", "Total Sales"])

for cell in region_ws[1]:
    cell.font = Font(bold=True, size=14)

region_totals = {}


for row in range(2, ws.max_row + 1):
    region = ws.cell(row=row, column=3).value
    total = ws.cell(row=row, column=6).value

    if region in region_totals:
        region_totals[region] += total
    else:
        region_totals[region] = total


for region, total in region_totals.items(): #skriver resultaten till Excel
    region_ws.append([region, total])


tab = Table(displayName="RegionTable", ref=f"A1:B{region_ws.max_row}")

style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

tab.tableStyleInfo = style
region_ws.add_table(tab)

chart = LineChart()
chart.title = "Total Sales per Region"
chart.y_axis.title = "Sales"
chart.x_axis.title = "Region"


data = Reference(region_ws, min_col=2, min_row=1, max_row=region_ws.max_row) #väljer vilken data som ska visas i diagrammet
categories = Reference(region_ws, min_col=1, min_row=2, max_row=region_ws.max_row)

chart.add_data(data, titles_from_data=True) #ägger till försäljningsdata i diagrammet
chart.set_categories(categories)


region_ws.add_chart(chart, "D2")


wb.save(output_path)
